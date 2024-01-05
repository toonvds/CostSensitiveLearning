import numpy as np
import matplotlib.pyplot as plt
from sklearn import metrics
from scipy.stats import spearmanr, combine_pvalues, friedmanchisquare, lognorm
from scikit_posthocs import posthoc_nemenyi_friedman
from tabulate import tabulate
from Orange.evaluation import compute_CD, graph_ranks
from hmeasure import h_score
import os
import baycomp
from openpyxl import workbook, load_workbook


# from rpy2.robjects.packages import importr
# import rpy2.robjects.numpy2ri
# hmeasure = importr('hmeasure')
# rpy2.robjects.numpy2ri.activate()

# Matplotlib settings for figures:
# plt.style.use('science')
# plt.rcParams.update({'font.size': 14})
# plt.rc('xtick', labelsize=12)
# plt.rc('ytick', labelsize=12)
# plt.rcParams['figure.figsize'] = (7, 6)
# plt.rcParams['figure.dpi'] = 250


def savings(cost_matrix, labels, predictions):
    cost_without = cost_without_algorithm(cost_matrix, labels)
    cost_with = cost_with_algorithm(cost_matrix, labels, predictions)
    savings = 1 - cost_with / cost_without

    return savings


def cost_with_algorithm(cost_matrix, labels, predictions):
    cost_tn = cost_matrix[:, 0, 0][np.logical_and(predictions == 0, labels == 0)].sum()
    cost_fn = cost_matrix[:, 0, 1][np.logical_and(predictions == 0, labels == 1)].sum()
    cost_fp = cost_matrix[:, 1, 0][np.logical_and(predictions == 1, labels == 0)].sum()
    cost_tp = cost_matrix[:, 1, 1][np.logical_and(predictions == 1, labels == 1)].sum()

    return sum((cost_tn, cost_fn, cost_fp, cost_tp))


def cost_without_algorithm(cost_matrix, labels):
    # Predict everything as the default class that leads to minimal cost
    # Also include cost of TP/TN!
    cost_neg = cost_matrix[:, 0, 0][labels == 0].sum() + cost_matrix[:, 0, 1][labels == 1].sum()
    cost_pos = cost_matrix[:, 1, 0][labels == 0].sum() + cost_matrix[:, 1, 1][labels == 1].sum()

    return min(cost_neg, cost_pos)


def rociv(labels, probabilities, costs):
    # Total cost per class
    pos_total = costs[labels == 1].sum()
    neg_total = costs[labels == 0].sum()

    # Sort predictions (1 to 0) and corresponding labels
    sorted_indices = np.argsort(probabilities)[::-1]
    costs_sorted = costs[sorted_indices]
    labels_sorted = labels[sorted_indices]
    # probabilities[sorted_indices][labels_sorted == 1]

    # Create ROCIV curve
    fp_costs = [0]
    tp_benefits = [0]

    benefits_cum = 0
    costs_cum = 0

    for i in range(len(labels)):
        if labels_sorted[i]:
            benefits_cum += costs_sorted[i]
        else:
            costs_cum += costs_sorted[i]

        fp_costs.append(costs_cum / neg_total)
        tp_benefits.append(benefits_cum / pos_total)

    # Area under curve
    auciv = metrics.auc(x=fp_costs, y=tp_benefits)
    # auciv = np.trapz(y=tp_benefits, x=fp_costs)

    return fp_costs, tp_benefits, auciv


def get_performance_metrics(evaluators, evaluation_matrices, i, index, cost_matrix, labels, probabilities, predictions,
                            info):
    if evaluators['traditional']:
        true_pos = (predictions * labels).sum()
        true_neg = ((1 - predictions) * (1 - labels)).sum()
        false_pos = (predictions * (1 - labels)).sum()
        false_neg = ((1 - predictions) * labels).sum()

        accuracy = (true_pos + true_neg) / len(labels)
        recall = true_pos / (true_pos + false_neg)
        # Make sure no division by 0!
        if (true_pos == 0) and (false_pos == 0):
            precision = 0
            print('\t\tWARNING: No positive predictions!')
        else:
            precision = true_pos / (true_pos + false_pos)
        if precision == 0:
            f1_score = 0
            print('\t\tWARNING: Precision = 0!')
        else:
            f1_score = 2 * (precision * recall) / (precision + recall)

        evaluation_matrices['traditional'][index, i] = np.array([accuracy, recall, precision, f1_score])

    if evaluators['ROC']:
        fpr, tpr, roc_thresholds = metrics.roc_curve(y_true=labels, y_score=probabilities)

        evaluation_matrices['ROC'][index, i] = np.array([fpr, tpr, roc_thresholds])

    if evaluators['AUC']:
        auc = metrics.roc_auc_score(y_true=labels, y_score=probabilities)

        evaluation_matrices['AUC'][index, i] = auc

    if evaluators['savings']:
        # To do: function - savings
        cost_without = cost_without_algorithm(cost_matrix, labels)
        cost_with = cost_with_algorithm(cost_matrix, labels, predictions)
        savings = 1 - cost_with / cost_without

        evaluation_matrices['savings'][index, i] = savings

    if evaluators['AEC']:
        expected_cost = labels * (probabilities * cost_matrix[:, 1, 1] + (1 - probabilities) * cost_matrix[:, 0, 1]) \
                        + (1 - labels) * (
                                probabilities * cost_matrix[:, 1, 0] + (1 - probabilities) * cost_matrix[:, 0, 0])

        aec = expected_cost.mean()

        evaluation_matrices['AEC'][index, i] = aec

    if evaluators['ROCIV']:
        misclass_costs = np.zeros(len(labels))
        misclass_costs[labels == 0] = cost_matrix[:, 1, 0][labels == 0]
        misclass_costs[labels == 1] = cost_matrix[:, 0, 1][labels == 1]
        fpcosts, tpbenefits, auciv = rociv(labels, probabilities, misclass_costs)

        evaluation_matrices['ROCIV'][index, i] = np.array([fpcosts, tpbenefits, auciv], dtype=object)

    if evaluators['H_measure']:
        # TODO:
        #   Takes approx 1 sec -> Do from scratch?
        #   https://github.com/canagnos/hmeasure-python/blob/master/mcp/mcp.py
        #   Specify cost distribution?
        #   Uses hmeasure, see https://github.com/cran/hmeasure/blob/master/R/library_metrics.R

        misclass_neg = cost_matrix[:, 1, 0][labels == 0]
        misclass_pos = cost_matrix[:, 0, 1][labels == 1]

        # Todo: explain this severity!
        severity = misclass_neg.mean() / misclass_pos.mean()

        # h = hmeasure.HMeasure(labels, probabilities[:, None], severity)[0][0][0]

        h = h_score(labels, probabilities, severity)

        evaluation_matrices['H_measure'][index, i] = h

    if evaluators['PR']:
        precision, recall, _ = metrics.precision_recall_curve(y_true=labels, probas_pred=probabilities)

        # AUC is not recommended here (see sklearn docs)
        # We will use Average Precision (AP)
        ap = metrics.average_precision_score(y_true=labels, y_score=probabilities)

        evaluation_matrices['PR'][index, i] = np.array([precision, recall, ap], dtype=object)

    if evaluators['PRIV']:
        misclass_costs = np.zeros(len(labels))
        misclass_costs[labels == 0] = cost_matrix[:, 1, 0][labels == 0]
        misclass_costs[labels == 1] = cost_matrix[:, 0, 1][labels == 1]

        precisioniv, recalliv, _ = metrics.precision_recall_curve(y_true=labels, probas_pred=probabilities,
                                                                  sample_weight=misclass_costs)

        # AUC is not recommended here (see sklearn docs)
        # We will use Average Precision (AP)
        apiv = metrics.average_precision_score(y_true=labels, y_score=probabilities, sample_weight=misclass_costs)
        # ap = metrics.auc(recall, precision)

        evaluation_matrices['PRIV'][index, i] = np.array([precisioniv, recalliv, apiv], dtype=object)

    if evaluators['rankings']:
        pos_probas = probabilities[labels == 1]
        # Get C_(0,1) (FN - misclassification cost of positive instances)
        misclass_costs_pos = cost_matrix[:, 0, 1][labels == 1]

        # Sort indices from high to low
        sorted_indices_probas = np.argsort(pos_probas)[::-1]
        prob_rankings = np.argsort(sorted_indices_probas)

        sorted_indices_amounts = np.argsort(misclass_costs_pos)[::-1]
        amount_rankings = np.argsort(sorted_indices_amounts)

        #  Compare rankings of probas with rankings of amounts for all positive instances
        spearman_test = spearmanr(prob_rankings, amount_rankings)

        evaluation_matrices['rankings'][index, i] = np.array([misclass_costs_pos[sorted_indices_probas], spearman_test],
                                                             dtype=object)

    if evaluators['brier']:
        brier = ((probabilities - labels) ** 2).mean()

        evaluation_matrices['brier'][index, i] = brier

    if evaluators['recall_overlap']:
        recalled = labels[labels == 1] * predictions[labels == 1]

        evaluation_matrices['recall_overlap'][index, i] = recalled

    if evaluators['recall_correlation']:
        pos_probas = probabilities[labels == 1]

        # Sort indices from high to low
        sorted_indices_probas = np.argsort(pos_probas)[::-1]
        prob_rankings = np.argsort(sorted_indices_probas)

        evaluation_matrices['recall_correlation'][index, i] = prob_rankings

    if evaluators['time']:
        evaluation_matrices['time'][index, i] = info['time']

    if evaluators['lambda1']:
        evaluation_matrices['lambda1'][index, i] = info['lambda1']

    if evaluators['lambda2']:
        evaluation_matrices['lambda2'][index, i] = info['lambda2']

    if evaluators['n_neurons']:
        evaluation_matrices['n_neurons'][index, i] = info['n_neurons']

    return evaluation_matrices


def evaluate_experiments(evaluators, methodologies, evaluation_matrices, directory, name):
    table_evaluation = []
    n_methodologies = sum(methodologies.values())

    names = []
    for key in methodologies.keys():
        if methodologies[key]:
            names.append(key)

    if evaluators['traditional']:

        table_traditional = [['Method', 'Accuracy', 'Recall', 'Precision', 'F1-score', 'AR', 'sd']]

        # Compute F1 rankings (- as higher is better)
        all_f1s = []
        for i in range(evaluation_matrices['traditional'].shape[0]):
            method_f1s = []
            for j in range(evaluation_matrices['traditional'][i].shape[0]):
                f1 = evaluation_matrices['traditional'][i][j][-1]
                method_f1s.append(f1)
            all_f1s.append(np.array(method_f1s))

        ranked_args = np.argsort(-np.array(all_f1s), axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = np.mean(rankings, axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize all per method
        index = 0
        for item, value in methodologies.items():
            if value:
                averages = evaluation_matrices['traditional'][index, :].mean()

                table_traditional.append([item, averages[0], averages[1], averages[2], averages[3],
                                          avg_rankings[index], sd_rankings[index]])

                index += 1

        print(tabulate(table_traditional, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_traditional)

        # Do tests if enough measurements are available (at least 3)
        if np.array(all_f1s).shape[1] > 2:
            friedchisq = friedmanchisquare(*np.transpose(all_f1s))
            print('\nF1 - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                # Post-hoc Nemenyi Friedman: Rows are blocks, columns are groups
                nemenyi = posthoc_nemenyi_friedman(np.array(all_f1s).T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['ROC']:
        index = 0
        # fig, ax = plt.subplots()
        # ax.set_title('ROC curve')
        # ax.set_xlabel('False positive rate')
        # ax.set_ylabel('True positive rate')
        # for item, value in methodologies.items():
        #     if value:
        #         # See https://scikit-learn.org/stable/auto_examples/model_selection/plot_roc_crossval.html
        #         tprs = []
        #         mean_fpr = np.linspace(0, 1, 100)
        #
        #         for i in range(evaluation_matrices['ROC'][index, :].shape[0]):
        #             fpr, tpr, _ = list(evaluation_matrices['ROC'][index, i])
        #             interp_tpr = np.interp(mean_fpr, fpr, tpr)
        #             interp_tpr[0] = 0.0
        #             tprs.append(interp_tpr)
        #
        #         mean_tpr = np.mean(tprs, axis=0)
        #         mean_tpr[-1] = 1.0
        #
        #         index += 1
        #
        #         ax.plot(mean_fpr, mean_tpr, label=item, lw=2, alpha=.8)
        #
        #         # std_tpr = np.std(tprs, axis=0)
        #         # tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
        #         # tprs_lower = np.maximum(mean_tpr - std_tpr, 0)
        #         # ax.fill_between(mean_fpr, tprs_lower, tprs_upper, color='grey', alpha=.2)
        #
        # ax.legend()
        # plt.savefig(str(directory + 'ROC.png'), bbox_inches='tight')
        # plt.show()

    if evaluators['AUC']:

        table_auc = [['Method', 'AUC', 'sd', 'AR', 'sd']]

        # Compute rankings (- as higher is better)
        ranked_args = (-evaluation_matrices['AUC']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_auc.append([item, evaluation_matrices['AUC'][index, :].mean(),
                                  np.sqrt(evaluation_matrices['AUC'][index, :].var()), avg_rankings[index],
                                  sd_rankings[index]])
                index += 1

        print(tabulate(table_auc, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_auc)

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['AUC'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['AUC'].T)
            print('\nAUC - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['AUC'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['savings']:

        table_savings = [['Method', 'Savings', 'sd', 'AR', 'sd']]

        # Compute rankings (- as higher is better)
        ranked_args = (-evaluation_matrices['savings']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        methods_used = []
        for item, value in methodologies.items():
            if value:
                methods_used.append(item)
                table_savings.append([item, evaluation_matrices['savings'][index, :].mean(),
                                      np.sqrt(evaluation_matrices['savings'][index, :].var()), avg_rankings[index],
                                      sd_rankings[index]])
                index += 1

        print(tabulate(table_savings, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_savings)

        # plt.xlabel('Methods')
        # plt.ylabel('Savings')
        # # plt.ylim(0, 1)
        # plt.boxplot(np.transpose(evaluation_matrices['savings']))
        # plt.xticks(np.arange(n_methodologies) + 1, methods_used)
        # plt.xticks(rotation=40)
        # plt.savefig(str(directory + 'savings_boxplot_' + name + '.png'), bbox_inches='tight')
        # plt.show()

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['savings'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['savings'].T)
            print('\nSavings - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['savings'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        if n_methodologies > 1:
            cd = compute_CD(avg_rankings, n=1, alpha='0.05', test="nemenyi")
            print(f'Critical difference: {np.round(cd, 4)}')
            graph_ranks(avg_rankings, names, cd=cd, width=9, textspace=3, lowv=1, highv=n_methodologies)
            # plt.show()

            # # Bayesian testing for all combinations:
            # print('Bayesian tests (ROPE) for savings (one vs one):')
            # for i in range(0, n_methodologies - 1):
            #     for j in range(i + 1, n_methodologies):
            #         print(str('\tComparing ' + names[i] + ' and ' + names[j]))
            #         probs = baycomp.two_on_single(evaluation_matrices['savings'][i], evaluation_matrices['savings'][j],
            #                                       plot=False, names=[names[i], names[j]])
            #         print(f'\t{probs}')
            #
            # print('Bayesian tests (ROPE) for rankings (one vs one):')
            # for i in range(0, n_methodologies - 1):
            #     for j in range(i + 1, n_methodologies):
            #         print(str('\tComparing ' + names[i] + ' and ' + names[j]))
            #         probs = baycomp.two_on_single(evaluation_matrices['savings'][i], evaluation_matrices['savings'][j],
            #                                       rope=1, plot=False, names=[names[i], names[j]])
            #         print(f'\t{probs}')
            #
            # # Bayesian testing multiple comparison:
            # # Not implemented yet
            # # print('Bayesian tests (ROPE) for savings (multiple comparisons):')

        print('_________________________________________________________________________')

    if evaluators['AEC']:

        table_aec = [['Method', 'AEC', 'sd', 'AR', 'sd']]

        # Compute rankings (lower is better)
        ranked_args = (evaluation_matrices['AEC']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        methods_used = []
        for item, value in methodologies.items():
            if value:
                methods_used.append(item)
                table_aec.append([item, evaluation_matrices['AEC'][index, :].mean(),
                                  np.sqrt(evaluation_matrices['AEC'][index, :].var()), avg_rankings[index],
                                  sd_rankings[index]])
                index += 1

        print(tabulate(table_aec, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_aec)

        # plt.xlabel('Methods')
        # plt.ylabel('AEC')
        # # plt.ylim(0, 1)
        # plt.boxplot(np.transpose(evaluation_matrices['AEC']))
        # plt.xticks(np.arange(n_methodologies) + 1, methods_used)
        # plt.xticks(rotation=40)
        # plt.savefig(str(directory + 'AEC_boxplot' + '.png'), bbox_inches='tight')
        # plt.show()

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['AEC'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['AEC'].T)
            print('\nSavings - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['AEC'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['ROCIV']:
        table_auciv = [['Method', 'AUCIV', 'sd', 'AR', 'sd']]

        index = 0
        # fig2, ax2 = plt.subplots()
        # ax2.set_title('ROCIV curve')
        # ax2.set_xlabel('False positive cost')
        # ax2.set_ylabel('True positive benefit')

        all_aucivs = []

        for item, value in methodologies.items():
            if value:
                # See https://scikit-learn.org/stable/auto_examples/model_selection/plot_roc_crossval.html
                tprs = []
                mean_fpr = np.linspace(0, 1, 100)
                aucivs = []

                for i in range(evaluation_matrices['ROCIV'][index, :].shape[0]):
                    fp_costs, tp_benefits, auciv = list(evaluation_matrices['ROCIV'][index, i])
                    interp_tpr = np.interp(mean_fpr, fp_costs, tp_benefits)
                    interp_tpr[0] = 0.0
                    tprs.append(interp_tpr)
                    aucivs.append(auciv)

                mean_tpr = np.mean(tprs, axis=0)
                mean_tpr[-1] = 1.0

                # ax2.plot(mean_fpr, mean_tpr, label=item, lw=2, alpha=.8)
                # std_tpr = np.std(tprs, axis=0)
                # tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
                # tprs_lower = np.maximum(mean_tpr - std_tpr, 0)
                # ax2.fill_between(mean_fpr, tprs_lower, tprs_upper, color='grey', alpha=.2)

                aucivs = np.array(aucivs)
                table_auciv.append([item, aucivs.mean(), np.sqrt(aucivs.var())])
                all_aucivs.append(aucivs)

                index += 1

        # ax2.legend()
        # plt.savefig(str(directory + 'ROCIV.png'), bbox_inches='tight')
        # plt.show()

        # Add rankings (higher is better)
        ranked_args = np.argsort(-np.array(all_aucivs), axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = np.mean(rankings, axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))
        for i in range(1, len(table_auciv)):
            table_auciv[i].append(avg_rankings[i - 1])
            table_auciv[i].append(sd_rankings[i - 1])

        print(tabulate(table_auciv, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_auciv)

        # Do tests if enough measurements are available (at least 3)
        if np.array(all_aucivs).shape[1] > 2:
            friedchisq = friedmanchisquare(*np.transpose(all_aucivs))
            print('\nAUCIV - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(np.array(all_aucivs).T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['H_measure']:

        table_H = [['Method', 'H_measure', 'sd', 'AR', 'sd']]

        # Compute rankings (- as higher is better)
        ranked_args = (-evaluation_matrices['H_measure']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_H.append([item, evaluation_matrices['H_measure'][index, :].mean(),
                                np.sqrt(evaluation_matrices['H_measure'][index, :].var()), avg_rankings[index],
                                sd_rankings[index]])
                index += 1

        print(tabulate(table_H, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_H)

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['H_measure'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['H_measure'].T)
            print('\nH-measure - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['H_measure'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['PR']:
        table_ap = [['Method', 'Avg Prec', 'sd', 'AR', 'sd']]

        index = 0
        # fig2, ax2 = plt.subplots()
        # ax2.set_title('PR curve')
        # ax2.set_xlabel('Recall')
        # ax2.set_ylabel('Precision')

        all_aps = []
        for item, value in methodologies.items():
            if value:
                # See https://scikit-learn.org/stable/auto_examples/model_selection/plot_roc_crossval.html
                precisions = []
                mean_recall = np.linspace(0, 1, 100)
                aps = []

                for i in range(evaluation_matrices['PR'][index, :].shape[0]):
                    precision, recall, ap = list(evaluation_matrices['PR'][index, i])

                    interp_precision = np.interp(mean_recall, recall[::-1], precision[::-1])
                    interp_precision[0] = 1
                    precisions.append(interp_precision)

                    aps.append(ap)

                mean_precision = np.mean(precisions, axis=0)
                mean_precision[-1] = 0

                # ax2.plot(mean_recall, mean_precision, label=item, lw=2, alpha=.8)
                # std_precision = np.std(precisions, axis=0)
                # precisions_upper = np.minimum(mean_precision + std_precision, 1)
                # precisions_lower = np.maximum(mean_precision - std_precision, 0)
                # ax2.fill_between(mean_recall, precisions_lower, precisions_upper, color='grey', alpha=.2)

                aps = np.array(aps)
                table_ap.append([item, aps.mean(), np.sqrt(aps.var())])

                all_aps.append(aps)

                index += 1

        # ax2.legend()
        # plt.savefig(str(directory + 'PR.png'), bbox_inches='tight')
        # plt.show()

        # Add rankings (higher is better)
        ranked_args = np.argsort(-np.array(all_aps), axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = np.mean(rankings, axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))
        for i in range(1, len(table_ap)):
            table_ap[i].append(avg_rankings[i - 1])
            table_ap[i].append(sd_rankings[i - 1])

        print(tabulate(table_ap, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_ap)

        # Do tests if enough measurements are available (at least 3)
        if np.array(all_aps).shape[1] > 2:
            friedchisq = friedmanchisquare(*np.transpose(all_aps))
            print('\nAP - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(np.array(all_aps).T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['PRIV']:
        table_apiv = [['Method', 'APIV', 'sd', 'AR', 'sd']]

        index = 0
        # fig2, ax2 = plt.subplots()
        # ax2.set_title('PRIV curve')
        # ax2.set_xlabel('Weighted recall')
        # ax2.set_ylabel('Weighted precision')

        all_apivs = []
        for item, value in methodologies.items():
            if value:
                # See https://scikit-learn.org/stable/auto_examples/model_selection/plot_roc_crossval.html
                precisions = []
                mean_recall = np.linspace(0, 1, 100)
                apivs = []

                for i in range(evaluation_matrices['PRIV'][index, :].shape[0]):
                    precision, recall, apiv = list(evaluation_matrices['PRIV'][index, i])

                    interp_precision = np.interp(mean_recall, recall[::-1], precision[::-1])
                    interp_precision[0] = 1
                    precisions.append(interp_precision)

                    apivs.append(apiv)

                mean_precision = np.mean(precisions, axis=0)
                mean_precision[-1] = 0

                # ax2.plot(mean_recall, mean_precision, label=item, lw=2, alpha=.8)
                # std_precision = np.std(precisions, axis=0)
                # precisions_upper = np.minimum(mean_precision + std_precision, 1)
                # precisions_lower = np.maximum(mean_precision - std_precision, 0)
                # ax2.fill_between(mean_recall, precisions_lower, precisions_upper, color='grey', alpha=.2)

                apivs = np.array(apivs)
                table_apiv.append([item, apivs.mean(), np.sqrt(apivs.var())])
                all_apivs.append(apivs)

                index += 1

        # ax2.legend()
        # plt.savefig(str(directory + 'PRIV.png'), bbox_inches='tight')
        # plt.show()

        # Add rankings (higher is better)
        ranked_args = np.argsort(-np.array(all_apivs), axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = np.mean(rankings, axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))
        for i in range(1, len(table_apiv)):
            table_apiv[i].append(avg_rankings[i - 1])
            table_apiv[i].append(sd_rankings[i - 1])

        print(tabulate(table_apiv, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4", ".4f")))
        table_evaluation.append(table_apiv)

        # Do tests if enough measurements are available (at least 3)
        if np.array(all_apivs).shape[1] > 2:
            friedchisq = friedmanchisquare(*np.transpose(all_apivs))
            print('\nAPIV - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(np.array(all_apivs).T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['rankings']:
        table_spearman = [['Method', 'Spearman corr', 'p-value', 'AR', 'sd']]

        index = 0
        # fig3, ax3 = plt.subplots()
        # ax3.set_title('Rankings')
        # ax3.set_xlabel('% positives included')
        # ax3.set_ylabel('% of total amounts')

        all_scs = []
        for item, value in methodologies.items():
            if value:
                # See https://scikit-learn.org/stable/auto_examples/model_selection/plot_roc_crossval.html
                total_cum_frac_amounts = []
                mean_included = np.linspace(0, 1, 100)
                spearman_corrs = []
                p_values = []

                for i in range(evaluation_matrices['rankings'][index, :].shape[0]):
                    ranked_amounts, spearman_test = list(evaluation_matrices['rankings'][index, i])
                    spearman_corr, p_value = spearman_test

                    cum_frac_amounts = np.cumsum(ranked_amounts) / ranked_amounts.sum()

                    interp_cum_frac_amounts = np.interp(mean_included,
                                                        np.arange(0, len(cum_frac_amounts)) / len(cum_frac_amounts),
                                                        cum_frac_amounts)
                    total_cum_frac_amounts.append(interp_cum_frac_amounts)

                    spearman_corrs.append(spearman_corr)
                    p_values.append(p_value)

                mean_cum_amounts = np.mean(total_cum_frac_amounts, axis=0)

                # ax3.plot(mean_cum_amounts, label=item, lw=2, alpha=.8)
                # std_precision = np.std(precisions, axis=0)
                # precisions_upper = np.minimum(mean_precision + std_precision, 1)
                # precisions_lower = np.maximum(mean_precision - std_precision, 0)
                # ax2.fill_between(mean_recall, precisions_lower, precisions_upper, color='grey', alpha=.2)

                spearman_corrs = np.array(spearman_corrs)
                # Combine p values with Fisher's method:
                p_value_total = combine_pvalues(p_values)[1]  # Take only the p-value (not statistic)
                table_spearman.append([item, spearman_corrs.mean(), p_value_total])

                all_scs.append(spearman_corrs)

                index += 1

        best_amounts = np.cumsum(np.sort(ranked_amounts)[::-1] / ranked_amounts.sum())
        best_amounts_100 = np.interp(mean_included, np.arange(0, len(best_amounts)) / len(best_amounts), best_amounts)
        # ax3.plot(best_amounts_100, linestyle='dashed', color='k', label='best possible')
        # ax3.plot(np.cumsum(np.ones(100) / 100), linestyle='dotted', color='k', label='random')

        # Add rankings (higher is better)
        ranked_args = np.argsort(-np.array(all_scs), axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = np.mean(rankings, axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))
        for i in range(1, len(table_spearman)):
            table_spearman[i].append(avg_rankings[i - 1])
            table_spearman[i].append(sd_rankings[i - 1])

        # ax3.legend(bbox_to_anchor=(1.04,1), prop={'size': 12})
        # fig3.set_size_inches((7, 4))

        # plt.savefig(str(directory + 'rankings.png'), bbox_inches='tight')
        # plt.show()

        print(tabulate(table_spearman, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table_spearman)

        # Do tests if enough measurements are available (at least 3)
        if np.array(all_scs).shape[1] > 2:
            friedchisq = friedmanchisquare(*np.transpose(all_scs))
            print('\nSpearman correlation - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(np.array(all_scs).T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['brier']:
        table_brier = [['Method', 'Brier score', 'sd', 'AR', 'sd']]

        # Compute rankings (lower is better)
        ranked_args = (evaluation_matrices['brier']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_brier.append([item, evaluation_matrices['brier'][index, :].mean(),
                                    np.sqrt(evaluation_matrices['brier'][index, :].var()), avg_rankings[index],
                                    sd_rankings[index]])
                index += 1

        print(tabulate(table_brier, headers="firstrow", floatfmt=("", ".6f", ".6f", ".4f", ".4f")))
        table_evaluation.append(table_brier)

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['brier'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['brier'].T)
            print('\nBrier score - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['brier'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['recall_overlap']:

        # Make table with only relevant methodologies
        table_recall_overlap = ['Recall overlaps']
        for meth in methodologies:
            if methodologies[meth]:
                table_recall_overlap.append(meth)
        table_recall_overlap = [table_recall_overlap]

        # Get recall overlap per experiment (fold/repeat)
        n_experiments = evaluation_matrices['recall_overlap'].shape[1]
        recall_overlaps = np.zeros((n_experiments, n_methodologies, n_methodologies))
        for n in range(n_experiments):
            for i in range(n_methodologies):
                for j in range(n_methodologies):
                    # if j > i:
                    #    break
                    recall_overlaps[n, i, j] = (
                            evaluation_matrices['recall_overlap'][i, n] == evaluation_matrices['recall_overlap'][
                        j, n]).mean()

        # Summarize over repeated experiments
        recall_overlaps = recall_overlaps.mean(axis=0)

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_recall_overlap.append([item] + list(recall_overlaps[index, :]))
                index += 1

        print(tabulate(table_recall_overlap, headers="firstrow", floatfmt=()))
        table_evaluation.append(table_recall_overlap)

        print('_________________________________________________________________________')

    if evaluators['recall_correlation']:

        # Make table with only relevant methodologies
        table_recall_correlations = ['Recall correlations']
        for meth in methodologies:
            if methodologies[meth]:
                table_recall_correlations.append(meth)
        table_recall_correlations = [table_recall_correlations]

        # Get recall correlation per experiment (fold/repeat)
        n_experiments = evaluation_matrices['recall_correlation'].shape[1]
        recall_correlations = np.zeros((n_experiments, n_methodologies, n_methodologies))
        for n in range(n_experiments):
            for i in range(n_methodologies):
                for j in range(n_methodologies):
                    # if j > i:
                    #    break
                    # Todo: Spearman's correlation R
                    spearman_corr = spearmanr(evaluation_matrices['recall_correlation'][i, n],
                                              evaluation_matrices['recall_correlation'][j, n])

                    recall_correlations[n, i, j] = spearman_corr[0]

        # Summarize over repeated experiments
        recall_correlations = recall_correlations.mean(axis=0)

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_recall_correlations.append([item] + list(recall_correlations[index, :]))
                index += 1

        print(tabulate(table_recall_correlations, headers="firstrow", floatfmt=()))
        table_evaluation.append(table_recall_correlations)

        print('_________________________________________________________________________')

    if evaluators['time']:

        table_time = [['Method', 'Time', 'sd', 'AR', 'sd']]

        # Compute rankings (lower is better)
        ranked_args = (evaluation_matrices['time']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_time.append([item, evaluation_matrices['time'][index, :].mean(),
                                   np.sqrt(evaluation_matrices['time'][index, :].var()), avg_rankings[index],
                                   sd_rankings[index]])
                index += 1

        print(tabulate(table_time, headers="firstrow", floatfmt=("", ".6f", ".6f", ".4f", ".4f")))
        table_evaluation.append(table_time)

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['time'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['time'].T)
            print('\nTime - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['time'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['lambda1']:

        table_lambda1 = [['Method', 'Lambda l1', 'sd']]

        # Compute rankings (lower is better)
        ranked_args = (evaluation_matrices['lambda1']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_lambda1.append([item, evaluation_matrices['lambda1'][index, :].mean(),
                                      np.sqrt(evaluation_matrices['lambda1'][index, :].var()), avg_rankings[index],
                                      sd_rankings[index]])
                index += 1

        print(tabulate(table_lambda1, headers="firstrow", floatfmt=("", ".6f", ".6f")))
        table_evaluation.append(table_lambda1)

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['lambda1'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['lambda1'].T)
            print('\nLambda1 - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['lambda1'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['lambda2']:

        table_lambda2 = [['Method', 'Lambda2', 'sd']]

        # Compute rankings (lower is better)
        ranked_args = (evaluation_matrices['lambda2']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_lambda2.append([item, evaluation_matrices['lambda2'][index, :].mean(),
                                      np.sqrt(evaluation_matrices['lambda2'][index, :].var()), avg_rankings[index],
                                      sd_rankings[index]])
                index += 1

        print(tabulate(table_lambda2, headers="firstrow", floatfmt=("", ".6f", ".6f")))
        table_evaluation.append(table_lambda2)

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['lambda2'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['lambda2'].T)
            print('\nLambda l2 - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['lambda2'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['n_neurons']:

        table_n_neurons = [['Method', 'Number of neurons', 'sd']]

        # Compute rankings (lower is better)
        ranked_args = (evaluation_matrices['n_neurons']).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table_n_neurons.append([item, evaluation_matrices['n_neurons'][index, :].mean(),
                                        np.sqrt(evaluation_matrices['n_neurons'][index, :].var()), avg_rankings[index],
                                        sd_rankings[index]])
                index += 1

        print(tabulate(table_n_neurons, headers="firstrow", floatfmt=("", ".6f", ".6f")))
        table_evaluation.append(table_n_neurons)

        # Do tests if enough measurements are available (at least 3)
        if evaluation_matrices['n_neurons'].shape[1] > 2:
            friedchisq = friedmanchisquare(*evaluation_matrices['n_neurons'].T)
            print('\nNumber of neurons - Friedman test')
            print('H0: Model performance follows the same distribution')
            print('\tChi-square:\t%.4f' % friedchisq[0])
            print('\tp-value:\t%.4f' % friedchisq[1])
            if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
                nemenyi = posthoc_nemenyi_friedman(evaluation_matrices['n_neurons'].T.astype(dtype=np.float32))
                print('\nNemenyi post hoc test:')
                print(nemenyi)

        print('_________________________________________________________________________')

    if evaluators['LaTeX']:
        top_row = ['Method']
        n_methods = len(table_evaluation[0]) - 1
        bottom_rows = [None] * n_methods
        # Print to LaTeX:
        for i in range(len(table_evaluation)):
            # Print tables separately
            print(tabulate(table_evaluation[i], headers="firstrow", floatfmt=".4f", tablefmt='latex'))
            for j in range(len(table_evaluation[i])):
                if j == 0:
                    top_row += table_evaluation[i][j][1:]
                else:  # j>0
                    if i == 0:
                        bottom_rows[j - 1] = [table_evaluation[i][j][0]]
                    bottom_rows[j - 1] += table_evaluation[i][j][1:]
        total_table = [top_row] + bottom_rows

        print('\n' + tabulate(total_table, headers="firstrow", floatfmt=".4f"))
        print('\n' + tabulate(total_table, headers="firstrow", floatfmt=".4f", tablefmt='latex'))

    # Add to summary file
    with open(str(directory + 'summary.txt'), 'a') as file:
        for i in table_evaluation:
            file.write(tabulate(i, floatfmt=".4f") + '\n')

        if evaluators['LaTeX']:
            file.write(tabulate(total_table, headers="firstrow", floatfmt=".4f", tablefmt='latex') + '\n')

        # Save evaluation matrices to txt file
        if not os.path.isdir(directory + 'Evaluation_matrices/'):
            os.mkdir(directory + 'Evaluation_matrices/')
        for key in evaluation_matrices.keys():
            if evaluators[key]:
                np.save(str(directory + 'Evaluation_matrices/' + 'eval_np_' + key + '_' + name + '.npy'),
                        evaluation_matrices[key])  # '.txt', fmt='%s' Use for np.savetxt

        # for f in os.listdir(directory + 'Evaluation_matrices'):
        #     if f == 'desktop.ini':
        #         continue
        #     arr = np.load(directory + 'Evaluation_matrices/' + f, allow_pickle=True)
        #     print(arr)


def get_performance_metrics_ranking(evaluators, evaluation_matrices, i, index, cost_matrix, labels, probabilities,
                                    info=None):
    if evaluators['spearman_rho']:
        spearman_rho = rankings_spearman_rho(probabilities, labels, cost_matrix)[0]

        evaluation_matrices['spearman_rho'][index, i] = spearman_rho

    if evaluators['top10']:
        top10 = sum(topKinstances(probabilities, labels, k=10))

        evaluation_matrices['top10'][index, i] = top10

    if evaluators['top100']:
        top100 = sum(topKinstances(probabilities, labels, k=100))

        evaluation_matrices['top100'][index, i] = top100

    if evaluators['top500']:
        top500 = sum(topKinstances(probabilities, labels))

        evaluation_matrices['top500'][index, i] = top500

    if evaluators['top1000']:
        top1000 = sum(topKinstances(probabilities, labels, k=1000))

        evaluation_matrices['top1000'][index, i] = top1000

    if evaluators['top100costs']:
        top100costs = topKcosts(probabilities, labels, cost_matrix, k=100)

        evaluation_matrices['top100costs'][index, i] = top100costs

    if evaluators['top500costs']:
        top500costs = topKcosts(probabilities, labels, cost_matrix, k=500)

        evaluation_matrices['top500costs'][index, i] = top500costs

    if evaluators['top1000costs']:
        top1000costs = topKcosts(probabilities, labels, cost_matrix, k=1000)

        evaluation_matrices['top1000costs'][index, i] = top1000costs

    if evaluators['top10profit']:
        top10profit = topKprofit(probabilities, labels, cost_matrix, k=10)

        evaluation_matrices['top10profit'][index, i] = top10profit

    if evaluators['top100profit']:
        top100profit = topKprofit(probabilities, labels, cost_matrix, k=100)

        evaluation_matrices['top100profit'][index, i] = top100profit

    if evaluators['top500profit']:
        top500profit = topKprofit(probabilities, labels, cost_matrix, k=500)

        evaluation_matrices['top500profit'][index, i] = top500profit

    if evaluators['top1000profit']:
        top1000profit = topKprofit(probabilities, labels, cost_matrix, k=1000)

        evaluation_matrices['top1000profit'][index, i] = top1000profit

    if evaluators['cumulative_recovered']:
        sorted_labels = [x for _, x in sorted(zip(probabilities, labels))][::-1]

        # Sort amounts:
        sorted_amounts = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 0, 1]))][::-1]

        sorted_recovered = [a * b for a, b in zip(sorted_labels, sorted_amounts)]

        evaluation_matrices['cumulative_recovered'][index, i] = sorted_recovered

    if evaluators['cumulative_profit']:
        # Get rankings:
        rankings = np.argsort(probabilities)[::-1]

        # Sort labels and cost matrix:
        sorted_labels = labels[rankings]
        sorted_cost_matrix = np.copy(cost_matrix[rankings, :, :])

        # sorted_labels = [x for _, x in sorted(zip(probabilities, labels))][::-1]
        #
        # sorted_cost_matrix = np.copy(cost_matrix)
        # sorted_cost_matrix[:, 0, 0] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 0, 0]))][::-1]
        # sorted_cost_matrix[:, 0, 1] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 0, 1]))][::-1]
        # sorted_cost_matrix[:, 1, 0] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 1, 0]))][::-1]
        # sorted_cost_matrix[:, 1, 1] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 1, 1]))][::-1]

        # Calculate recovered profit (positive instances) and incurred costs (negative instances)
        recovered = np.array(sorted_labels) * (sorted_cost_matrix[:, 0, 1] - sorted_cost_matrix[:, 1, 1])
        costs = (1 - np.array(sorted_labels)) * (sorted_cost_matrix[:, 0, 0] - sorted_cost_matrix[:, 1, 0])

        actual = recovered + costs

        # Calculate best possible profit:

        ideal_recovered = np.array(sorted(recovered)[::-1])
        ideal_costs = np.array(sorted(costs)[::-1])

        ideal = ideal_recovered + ideal_costs

        evaluation_matrices['cumulative_profit'][index, i] = [actual, ideal]

    if evaluators['AP']:
        ap = metrics.label_ranking_average_precision_score(np.stack((labels, 1 - labels), axis=1),
                                                           np.stack((probabilities, 1 - probabilities), axis=1))

        evaluation_matrices['AP'][index, i] = ap

    if evaluators['expected_precision']:
        expected_prec = expected_precision(probabilities, labels)

        evaluation_matrices['expected_precision'][index, i] = expected_prec

    if evaluators['expected_profit']:
        expected_prof = expected_profit(probabilities, labels, cost_matrix)

        evaluation_matrices['expected_profit'][index, i] = expected_prof

    # if evaluators['CE']:
    #     ce = metrics.coverage_error(np.stack((labels, 1 - labels), axis=1),
    #                     np.stack((probabilities, 1 - probabilities), axis=1))

    return evaluation_matrices


# Spearman rho (predictions / amounts)
def rankings_spearman_rho(probabilities, labels, cost_matrix):
    pos_probas = probabilities[labels == 1]
    # Get C_(0,1) (FN - misclassification cost of positive instances)
    misclass_costs_pos = cost_matrix[:, 0, 1][labels == 1]

    # Sort indices from high to low
    sorted_indices_probas = np.argsort(pos_probas)[::-1]
    prob_rankings = np.argsort(sorted_indices_probas)

    sorted_indices_amounts = np.argsort(misclass_costs_pos)[::-1]
    amount_rankings = np.argsort(sorted_indices_amounts)

    #  Compare rankings of probas with rankings of amounts for all positive instances
    spearman_test = spearmanr(prob_rankings, amount_rankings)

    return spearman_test


# top k predicted labels
def topKinstances(probabilities, labels, k=500):
    sorted_labels = [x for _, x in sorted(zip(probabilities, labels))][::-1]

    return sorted_labels[0:k]


# top k predicted labels * their cost of being classified as positive
def topKcosts(probabilities, labels, cost_matrix, k=500):
    sorted_labels = [x for _, x in sorted(zip(probabilities, labels))][::-1]

    # Sort cost matrix:
    sorted_cost_matrix = cost_matrix
    # sorted_cost_matrix[:, 0, 0] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 0, 0]))][::-1]
    # sorted_cost_matrix[:, 0, 1] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 0, 1]))][::-1]
    sorted_cost_matrix[:, 1, 0] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 1, 0]))][::-1]
    sorted_cost_matrix[:, 1, 1] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 1, 1]))][::-1]

    costs = np.array(sorted_labels) * sorted_cost_matrix[:, 1, 1] + \
            (1 - np.array(sorted_labels)) * sorted_cost_matrix[:, 1, 0]

    return costs[0:k].sum()


# top k predicted labels * their amounts
def topKrecovered(probabilities, labels, amounts, k=500):
    sorted_labels = [x for _, x in sorted(zip(probabilities, labels))][::-1]

    # Sort amounts:
    sorted_amounts = [x for _, x in sorted(zip(probabilities, amounts))][::-1]

    # costs = np.array(sorted_labels) * sorted_cost_matrix[:, 1, 1] + (1-np.array(sorted_labels)) * np.zeros((len(labels)))

    recovered = np.array(sorted_amounts[0:k]) * np.array(sorted_labels[0:k])

    return recovered.sum()


# top k predicted labels * their amounts
def topKprofit(probabilities, labels, cost_matrix, k=500):
    # Sort labels and cost matrix:
    # sorted_labels = [x for _, x in sorted(zip(probabilities, labels))][::-1]
    #
    # sorted_cost_matrix = np.copy(cost_matrix)
    # sorted_cost_matrix[:, 0, 0] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 0, 0]))][::-1]
    # sorted_cost_matrix[:, 0, 1] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 0, 1]))][::-1]
    # sorted_cost_matrix[:, 1, 0] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 1, 0]))][::-1]
    # sorted_cost_matrix[:, 1, 1] = [x for _, x in sorted(zip(probabilities, cost_matrix[:, 1, 1]))][::-1]

    # Get rankings:
    rankings = np.argsort(probabilities)[::-1]

    # Sort labels and cost matrix:
    sorted_labels = labels[rankings]
    sorted_cost_matrix = np.copy(cost_matrix[rankings, :, :])

    # Calculate recovered profit (positive instances) and incurred costs (negative instances)
    recovered = np.array(sorted_labels) * (sorted_cost_matrix[:, 0, 1] - sorted_cost_matrix[:, 1, 1])

    costs = (1 - np.array(sorted_labels)) * (sorted_cost_matrix[:, 0, 0] - sorted_cost_matrix[:, 1, 0])

    actual = recovered + costs

    # Calculate ideal:
    ideal_recovered = np.array(sorted(recovered)[::-1])
    ideal_costs = np.array(sorted(costs)[::-1])

    ideal = ideal_recovered + ideal_costs

    if ideal[0:k].sum() < 0:
        print('Ideal smaller than 0!')

    return actual[0:k].sum() / ideal[0:k].sum()


# Expected precision given capacity probability distribution
def expected_precision(probabilities, labels):
    # Get rankings:
    rankings = np.argsort(probabilities)[::-1]

    # Sort labels and cost matrix:
    actual = labels[rankings]

    # Calculate ideal:
    ideal = np.array(sorted(labels)[::-1])

    topKprecisions = [actual[0:k].sum() / ideal[0:k].sum() for k in range(1, len(labels) + 1)]

    indices = np.cumsum(np.ones(len(labels)))
    probabilities = lognorm.pdf(x=indices, s=1, loc=0, scale=100)

    return sum(topKprecisions*probabilities)


# Expected profit given capacity probability distribution
def expected_profit(probabilities, labels, cost_matrix):
    # Get rankings:
    rankings = np.argsort(probabilities)[::-1]

    # Sort labels and cost matrix:
    sorted_labels = labels[rankings]
    sorted_cost_matrix = np.copy(cost_matrix[rankings, :, :])

    # Calculate recovered profit (positive instances) and incurred costs (negative instances)
    recovered = np.array(sorted_labels) * (sorted_cost_matrix[:, 0, 1] - sorted_cost_matrix[:, 1, 1])

    costs = (1 - np.array(sorted_labels)) * (sorted_cost_matrix[:, 0, 0] - sorted_cost_matrix[:, 1, 0])

    actual = recovered + costs

    # Calculate ideal:
    ideal_recovered = np.array(sorted(recovered)[::-1])
    ideal_costs = np.array(sorted(costs)[::-1])

    ideal = ideal_recovered + ideal_costs

    topKprofits = [actual[0:k].sum() / ideal[0:k].sum() for k in range(1, len(labels) + 1)]

    indices = np.cumsum(np.ones(len(labels)))
    probabilities = lognorm.pdf(x=indices, s=1, loc=0, scale=100)

    return sum(topKprofits * probabilities)


def evaluate_ranking_experiments(evaluators, methodologies, evaluation_matrices, directory, name):
    table_evaluation = []
    n_methodologies = sum(methodologies.values())

    names = []
    for key in methodologies.keys():
        if methodologies[key]:
            names.append(key)

    # Summarize and print metrics:
    for key in evaluators.keys():
        if evaluators[key]:
            evaluate_metric(name=key, evaluation_matrix=evaluation_matrices[key],
                            table_evaluation=table_evaluation, methodologies=methodologies)

    # Write to Excel:
    # wb = load_workbook(str(directory + 'CS Learning to rank - exploration results.xlsx'))
    # sheet = wb['Sheet6']
    #
    # start_row = 4
    # dataset_column = 2 + 3  # Change (2 + 1 to 2 + 10)
    #
    # for table in table_evaluation:
    #     for model in range(1, len(table)):
    #         metric = table[model][1]
    #         sheet.cell(row=start_row + model, column=dataset_column).value = metric
    #
    #         ranking = table[model][3]
    #         sheet.cell(row=start_row + model, column=dataset_column + 14).value = ranking
    #
    #     start_row += 7  # Adjust depending on number of models (3 + # models)
    #
    # wb.save(str(directory + 'CS Learning to rank - exploration results.xlsx'))

    # Add to summary file
    with open(str(directory), 'a') as file:
        for i in table_evaluation:
            file.write(tabulate(i, floatfmt=".4f") + '\n')

        # if evaluators['LaTeX']:
        #     print('Not yet implemented')
        # file.write(tabulate(total_table, headers="firstrow", floatfmt=".4f", tablefmt='latex') + '\n')

        # Save evaluation matrices to txt file
        # if not os.path.isdir(directory + 'Evaluation_matrices/'):
        #     os.mkdir(directory + 'Evaluation_matrices/')
        # for key in evaluation_matrices.keys():
        #     if evaluators[key]:
        #         np.save(str(directory + 'Evaluation_matrices/' + 'eval_np_' + key + '_' + name + '.npy'),
        #                 evaluation_matrices[key])


def evaluate_metric(name, evaluation_matrix, table_evaluation, methodologies):
    print('\nMetric: ' + name)

    table = [['Method', name, 'sd', 'AR', 'sd']]

    if name == 'cumulative_recovered' or name == 'cumulative_profit':
        fig, ax = plt.subplots(dpi=500)
        if name == 'cumulative_recovered':
            ax.set_title('Cumulative amount recovered')
            ax.set_ylabel('Amount recovered [%]')
        elif name == 'cumulative_profit':
            ax.set_title('Cumulative profit')
            ax.set_ylabel('Profit')

            all_aucs_ideal = []

        ax.set_xlabel('Instances [%]')
        all_aucs = []

        index = 0
        for item, value in methodologies.items():
            if value:
                # See https://scikit-learn.org/stable/auto_examples/model_selection/plot_roc_crossval.html
                total_cum_frac_amounts = []
                mean_included = np.linspace(0, 1, 100)
                aucs = []

                if name == 'cumulative_profit':
                    total_cum_frac_amounts_ideal = []
                    aucs_ideal = []

                for i in range(evaluation_matrix.shape[1]):
                    if name == 'cumulative_recovered':
                        sorted_recovered = evaluation_matrix[index, i]

                        cum_frac_amounts = np.cumsum(sorted_recovered) / sum(sorted_recovered)

                        interp_cum_frac_amounts = np.interp(mean_included,
                                                            np.arange(0, len(cum_frac_amounts)) / len(cum_frac_amounts),
                                                            cum_frac_amounts)
                        total_cum_frac_amounts.append(interp_cum_frac_amounts)

                        aucs.append(metrics.auc(mean_included, interp_cum_frac_amounts))
                    elif name == 'cumulative_profit':
                        actual, ideal = evaluation_matrix[index, i]

                        # Get cumulative profit of model:
                        cum_frac_amounts = np.cumsum(actual, dtype=np.float64)
                        interp_cum_frac_amounts = np.interp(mean_included,
                                                            np.arange(0, len(cum_frac_amounts)) / len(cum_frac_amounts),
                                                            cum_frac_amounts)

                        total_cum_frac_amounts.append(interp_cum_frac_amounts)

                        # Calculate ideal:
                        cum_frac_amounts_ideal = np.cumsum(ideal)
                        interp_cum_frac_amounts_ideal = np.interp(mean_included,
                                                                  np.arange(0, len(cum_frac_amounts_ideal)) / len(
                                                                      cum_frac_amounts_ideal),
                                                                  cum_frac_amounts_ideal)
                        total_cum_frac_amounts_ideal.append(interp_cum_frac_amounts_ideal)

                        # Calculate auc:
                        aucs.append(metrics.auc(mean_included, interp_cum_frac_amounts))
                        aucs_ideal.append(metrics.auc(mean_included, interp_cum_frac_amounts_ideal))

                mean_cum_amounts = np.mean(total_cum_frac_amounts, axis=0)

                ax.plot(mean_cum_amounts, label=item, lw=2, alpha=.8)

                # std_precision = np.std(precisions, axis=0)
                # precisions_upper = np.minimum(mean_precision + std_precision, 1)
                # precisions_lower = np.maximum(mean_precision - std_precision, 0)
                # ax2.fill_between(mean_recall, precisions_lower, precisions_upper, color='grey', alpha=.2)

                all_aucs.append(aucs)
                if name == 'cumulative_profit':
                    all_aucs_ideal.append(aucs_ideal)

                index += 1

        # Add best possible/ideal:
        if name == 'cumulative_recovered':
            best_amounts = np.cumsum(np.sort(sorted_recovered)[::-1]) / sum(sorted_recovered)
            best_amounts_100 = np.interp(mean_included, np.arange(0, len(best_amounts)) / len(best_amounts),
                                         best_amounts)
            ax.plot(best_amounts_100, linestyle='dashed', color='k', label='best possible')
            ax.plot(np.cumsum(np.ones(100) / 100), linestyle='dotted', color='k', label='random')
        if name == 'cumulative_profit':
            ideal_profits = np.mean(total_cum_frac_amounts_ideal, axis=0)
            ax.plot(ideal_profits, linestyle='dashed', color='k', label='best possible')
            random_profits = np.interp(mean_included, [0, 1], [0, ideal_profits[-1]])
            ax.plot(random_profits, linestyle='dotted', color='k', label='random')

        ax.legend(bbox_to_anchor=(1.04, 1), prop={'size': 12})
        fig.set_size_inches((7, 4))

        # plt.savefig(str(directory + 'rankings.png'), bbox_inches='tight')
        # plt.show()

        # Compute rankings (- as higher is better)
        all_aucs = np.array(all_aucs)
        if name == 'cumulative_profit':
            all_aucs_ideal = np.array(all_aucs_ideal)

            # Calculate AUCS
            # See https://github.com/maks-sh/scikit-uplift/blob/c9dd56aa0277e81ef7c4be62bf2fd33432e46f36/sklift/metrics/metrics.py#L323
            auc_baseline = metrics.auc(mean_included, random_profits)
            # auc_ideal = metrics.auc(mean_included, ideal_profits) - auc_baseline
            all_aucs = (all_aucs - auc_baseline) / (all_aucs_ideal - auc_baseline)

        ranked_args = (-all_aucs).argsort(axis=0)
        rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
        rankings = rankings + 1
        avg_rankings = rankings.mean(axis=1)
        sd_rankings = np.sqrt(rankings.var(axis=1))

        # Summarize per method
        index = 0
        for item, value in methodologies.items():
            if value:
                table.append([item, all_aucs[index, :].mean(),
                              np.sqrt(all_aucs[index, :].var()), avg_rankings[index],
                              sd_rankings[index]])
                index += 1

        # Print
        print(tabulate(table, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
        table_evaluation.append(table)

        # # Do tests if enough measurements are available (at least 3)
        # if np.array(all_aucs).shape[1] > 2:
        #     friedchisq = friedmanchisquare(*np.transpose(all_aucs))
        #     print('\nF1 - Friedman test')
        #     print('H0: Model performance follows the same distribution')
        #     print('\tChi-square:\t%.4f' % friedchisq[0])
        #     print('\tp-value:\t%.4f' % friedchisq[1])
        #     if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
        #         # Post-hoc Nemenyi Friedman: Rows are blocks, columns are groups
        #         nemenyi = posthoc_nemenyi_friedman(np.array(all_aucs).T.astype(dtype=np.float32))
        #         print('\nNemenyi post hoc test:')
        #         print(nemenyi)

        return  # Stop function here

    # Compute rankings (- as higher is better)
    ranked_args = (-evaluation_matrix).argsort(axis=0)
    rankings = np.arange(len(ranked_args))[ranked_args.argsort(axis=0)]
    rankings = rankings + 1
    avg_rankings = rankings.mean(axis=1)
    sd_rankings = np.sqrt(rankings.var(axis=1))

    # Summarize per method
    index = 0
    for item, value in methodologies.items():
        if value:
            table.append([item, evaluation_matrix[index, :].mean(),
                          np.sqrt(evaluation_matrix[index, :].var()), avg_rankings[index],
                          sd_rankings[index]])
            index += 1

    # Print
    print(tabulate(table, headers="firstrow", floatfmt=("", ".4f", ".4f", ".4f", ".4f")))
    table_evaluation.append(table)

    # # Do tests if enough measurements are available (at least 3)
    # if np.array(evaluation_matrix).shape[1] > 2:
    #     friedchisq = friedmanchisquare(*np.transpose(evaluation_matrix))
    #     print('\nF1 - Friedman test')
    #     print('H0: Model performance follows the same distribution')
    #     print('\tChi-square:\t%.4f' % friedchisq[0])
    #     print('\tp-value:\t%.4f' % friedchisq[1])
    #     if friedchisq[1] < 0.05:  # If p-value is significant, do Nemenyi post hoc test
    #         # Post-hoc Nemenyi Friedman: Rows are blocks, columns are groups
    #         nemenyi = posthoc_nemenyi_friedman(np.array(evaluation_matrix).T.astype(dtype=np.float32))
    #         print('\nNemenyi post hoc test:')
    #         print(nemenyi)

    print('_________________________________________________________________________')
    print('_________________________________________________________________________')

